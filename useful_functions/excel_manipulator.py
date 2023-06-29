from openpyxl import load_workbook
from openpyxl.worksheet import worksheet as wss
from threading import Thread
from cbt_app.models import Question,Choice,Topic,UploadTitle



def xl2db(user,url,course,title):
    wb = load_workbook(url)
    ws = wb.active
    print(f'reading from the file: {ws["A2"].value}')

    # create uploadTitle db
    ut = UploadTitle.objects.create(examiner=user, title =title)

    
    for row in wss.Worksheet.iter_rows(ws,min_row=3,max_col=6,max_row=102):
        for cell in row:
            # if cell.value == None:
            #     print('question has finished')
            #     break
            print(cell.value, cell, cell.col_idx)
        
            # topic col
            if cell.col_idx ==1:
                # t=Topic.objects.create(name=cell.value)
                global t
                if cell.value == None:
                    print('no topic identified')
                    t=Topic.objects.get_or_create(name='no topic',courses = course)[0]
                else:
                    t=Topic.objects.get_or_create(name=cell.value,courses = course)[0]
            # question text col
            elif cell.col_idx ==2:
                # stop operation if question col is empty
                if cell.value == None:
                    print('question has finished')
                    return 1
                print(f'this is t in question block: {t}')
                global q
                q = Question.objects.get_or_create(content=cell.value)[0]
                q.upload_title.add(ut)
                print('title added to question id {q.id}') 
                # check if topic does not exist then create it
                if not q.topic.all():   
                    try:
                        q.topic.add(t)
                        print('topic has paired with question')
                    except:
                        print('couldnt pair question to topic')
            # options col
            elif cell.col_idx==3:
                print(f'now in option 1')
                if cell.value:
                    c1 = Choice.objects.update_or_create(content=cell.value,question=q)
                    print(f'option 1 not empty, executed')
                    
            elif cell.col_idx==4:
                print(f'now in option 2')
                if cell.value:
                    print('option 2 not empty, executed')
                    c2 = Choice.objects.update_or_create(content=cell.value,question=q)
            elif cell.col_idx==5:
                print(f'now in option 3')
                if cell.value:
                    print('option 3 not empty, executed')
                    c3 = Choice.objects.update_or_create(content=cell.value,question=q)
            elif cell.col_idx==6:
                print(f'now in option anser')
                if cell.value:
                    print('option answer not empty, executed')
                    a = Choice.objects.update_or_create(content=cell.value,question=q, is_answer=True)
            