from django.contrib.auth import get_user_model,get_user

# quiz form
"""
<!-- <form action="{% url 'cbt_app:quiz_progress' %}" method="post">
        {% csrf_token %}
        
        {% for question in page_obj %}
            <legend>
                <h3>{{forloop.counter}}. {{question.content}} </h3>
            </legend>
            {% if error_message %}
            <p><strong>{{ error_message }}</strong></p>
            {% endif %}
            <input type="hidden" value="{{question.id}}" name="question{{question.id}}">
            <fieldset>
                {% for choice in question.choice_set.all %}
                    <input  type="radio" name="choice{{question.id}}" value="{{choice.id}}" required> 
                    <label for="choice{{question.id}}">{{choice.content}} </label> <br>
                {% endfor %}
            </fieldset>
        {% endfor %}
        <input type="submit" value="submit" style="margin-top: 2em;">
    </form> -->

    <div class="pagination"> <span class="step-links">
        {% if page_obj.has_previous %}
        <a href="?page=1">&laquo; first</a> <a href="?page={{ page_obj.previous_page_number }}">previous</a>
        {% endif %}
        <span class="current">
        Page {{ page_obj.number }} of {{ page_obj.paginator.num_pages }}.
        </span>
        {% if page_obj.has_next %}
        <a href="?page={{ page_obj.next_page_number }}">next</a> <a href="?page={{ page_obj.paginator.num_pages }}">last &raquo;</a>
        {% endif %}
        </span>
    </div>

    <!-- without pagination -->
    {% for question in contents %}
    <!-- <div class="question">
        <h3>{{question.content}}</h3>

        <div class="choice">
            {% for choice in question.choice_set.all %}
            <input type="radio" id="{{choice.id}}">{{choice}} <br>
            {% endfor %}
        </div>
    </div> -->
    {% endfor %}

    <!-- when using formset -->
    <!-- {% for form in formset %}
        {{form.as_p}}
        <p>hello</p>
    {% endfor %} -->
"""

# views.py
"""
# @login_required
# def take_quiz(request,quiz_id):
#     quiz = Quiz.objects.get(id=quiz_id)
#     questions = Question.objects.filter(quiz__id = quiz_id).order_by('?')[0:3]
#     paginator = Paginator(questions, 6)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     context = {
#         'questions':questions,
#         'page_obj':page_obj
#     }
#     return render(request, 'quiz.html', context)

"""

# javascript request (get) data async: frontend
"""
x = await fetch('http://127.0.0.1:8000/quiz/api/')
y = await x.json()
console.log(y)
"""

"""
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet as wss
from threading import Thread
from cbt_app.models import Question,Choice,Topic

def xl2db(url,course,title):
    wb = load_workbook(url)
    ws = wb.active
    print(f'reading from the file: {ws["A2"].value}')
    inner = 0
    outer = 0
    for row in wss.Worksheet.iter_rows(ws,min_row=3,max_col=6,max_row=100):
        print(f'this is outer counter {outer}')
        # print(f'outer values{row[outer].value, row[outer], row[outer].col_idx}')

        # if cell is empty or no question
        # if row[outer].value == None:
        #     print('question has finished')
        #     break
        
        for cell in row:
            print(f'this inner counter {inner}')
            inner = inner +1
            # topic col
            if cell.col_idx ==1:
                # t=Topic.objects.create(name=cell.value)
                global t
                if cell.value == None:
                    print('no topic identified')
                    t=Topic.objects.get_or_create(name='no topic',courses = course)[0]
                else:
                    t=Topic.objects.get_or_create(name=cell.value,courses = course)[0]
            
            # question text col
            elif cell.col_idx ==2:
                # stop operation if question col is empty
                if cell.value == None:
                    print('no question found, exiting...')
                    return 1

                # if question already in db, move to another question
                try: 
                    q =Question.objects.get(content=cell.value)
                    print(f'question already in db... now adding title and questionid ={q.id}')
                    q.upload_title=title
                    q.save()
                except:
                    print(f'question does not exit, creating question')
                    print(cell.value, cell, cell.col_idx)
                    break
                    # print(f'this is t in question block: {t}')
                    global quest
                    # quest = Question.objects.get_or_create(content=cell.value,upload_title=title)[0]
                    quest =Question.objects.create(content=cell.value,upload_title=title)
                    print(f'this is quest id = {quest.id}')

                    try:
                        quest.topic.add(t)
                        print('topic has paired with question')
                    except:
                        print('couldnt pair question to topic')
            # options col
            elif cell.col_idx==3:
                if cell.value:
                    c1 = Choice.objects.create(content=cell.value,question=quest)
                    
            elif cell.col_idx==4:
                if cell.value:
                    c2 = Choice.objects.create(content=cell.value,question=quest)
            elif cell.col_idx==5:
                if cell.value:
                    c3 = Choice.objects.create(content=cell.value,question=quest)
            elif cell.col_idx==6:
                if cell.value:
                    a = Choice.objects.create(content=cell.value,question=quest, is_answer=True)
            
        outer = outer +1
            
"""

# timer controled duration db update
# function updateDb(){
#             // let data = {'time':timeLeft, 'sitting': {{sitting.id}} }
#             setInterval(function(){
#                 makeRequest("{% url 'cbt_app:time_update' sitting.quiz.id %}",'post',{'time':timeLeft/60, 'sitting': {{sitting.id}} })
#             }, 6000)
#         }


# loop through dict in template django
"""
{% for x,y in quiz_pair.items %}
            <tr>
                
                <td><a href="{% url 'cbt_app:result_list' x.id %}">{{x}}</a></td>
                <td>{{x.course}}</td>
                <td>{{x.level}}</td>
                <td>{{x.start_time}}</td>
                <td>{{x.is_available}}</td>
                <td>{{y}}</td>
            </tr>
            {% endfor %}
"""

# used to update duration in db through cmd
"""
for x in s: 
        try:                                            
            t = x.end_time - x.start_time
            t1 = t.total_seconds() / 60
            x.duration = round(t1,2)
            x.save()
            print(f'updated succesfully, duration = {x.duration}')
        except:
            print('endtime or start time null')
            continue
"""

#  update countdown time to db using event listener at regualar interval
"""
 function updateDb(){
            let data = {'time':timeLeft, 'sitting': {{sitting.id}} }
            setInterval(function()
            {
                makeRequest(
                    "{% url 'cbt_app:time_update' sitting.quiz.id %}",
                    'post',
                    {'time':timeLeft, 
                    'sitting': {{sitting.id}} }
                )
            }, 6000)
        }

        document.addEventListener(onload,updateDb())
"""